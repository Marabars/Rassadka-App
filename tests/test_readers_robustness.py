"""Tests for parser robustness with real-world file variations."""
from __future__ import annotations

import datetime
import io
import tempfile
from pathlib import Path

import openpyxl
import pytest

from app.utils.excel_utils import collect_date_columns, find_header_row, _try_parse_date

DATE_A = datetime.date(2026, 5, 4)
DATE_B = datetime.date(2026, 5, 5)
DATE_C = datetime.date(2026, 5, 6)


# ── excel_utils ──────────────────────────────────────────────────────────────

def test_find_header_row_first_row():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = "ФИО"
    assert find_header_row(ws, 1, "ФИО") == 1


def test_find_header_row_deep():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=25, column=2).value = "ФИО"
    # Default max_scan_rows=30 should find it
    assert find_header_row(ws, 2, "ФИО", max_scan_rows=30) == 25


def test_find_header_row_not_found_returns_none():
    wb = openpyxl.Workbook()
    ws = wb.active
    assert find_header_row(ws, 1, "ФИО") is None


def test_collect_date_columns_string_dates():
    """Dates stored as strings (e.g. from CSV-imported sheets) must be parsed."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = "04.05.2026"
    ws.cell(row=1, column=2).value = "2026-05-05"
    date_cols = collect_date_columns(ws, header_row=1, first_data_col=1)
    assert DATE_A in date_cols
    assert DATE_B in date_cols


def test_collect_date_columns_datetime_objects():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = datetime.datetime(2026, 5, 4)
    date_cols = collect_date_columns(ws, header_row=1, first_data_col=1)
    assert DATE_A in date_cols


def test_try_parse_date_formats():
    assert _try_parse_date("04.05.2026") == DATE_A
    assert _try_parse_date("2026-05-04") == DATE_A
    assert _try_parse_date("04/05/2026") == DATE_A
    assert _try_parse_date("not-a-date") is None


# ── choices_reader (blank rows in middle of list) ─────────────────────────────

def _make_choices_wb(rows: list[tuple]) -> openpyxl.Workbook:
    """Build an in-memory choices workbook. rows = list of (name, date, status)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2026"
    # header
    ws.cell(row=1, column=1).value = "ФИО"
    ws.cell(row=1, column=2).value = datetime.datetime(2026, 5, 4)
    for i, (name, status) in enumerate(rows, start=2):
        if name is not None:
            ws.cell(row=i, column=1).value = name
        ws.cell(row=i, column=2).value = status
    return wb


def _wb_to_path(wb: openpyxl.Workbook) -> Path:
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb.save(tmp.name)
    return Path(tmp.name)


STATUS_MAPPING = {
    "office": ["офис"],
    "remote": ["удаленно"],
    "vacation": ["отпуск"],
    "day_off": ["выходной"],
}


def test_choices_reader_skips_blank_rows_in_middle():
    from app.readers.choices_reader import read_choices
    rows = [("Alice", "офис"), (None, None), ("Bob", "удаленно")]
    path = _wb_to_path(_make_choices_wb(rows))
    choices, _ = read_choices(path, "2026", STATUS_MAPPING)
    names = {c.employee_name for c in choices}
    assert "Alice" in names
    assert "Bob" in names


def test_choices_reader_stops_after_many_consecutive_blanks():
    from app.readers.choices_reader import read_choices
    rows = [("Alice", "офис"), (None, None), (None, None), (None, None), ("Ghost", "офис")]
    path = _wb_to_path(_make_choices_wb(rows))
    choices, _ = read_choices(path, "2026", STATUS_MAPPING)
    names = {c.employee_name for c in choices}
    assert "Alice" in names
    assert "Ghost" not in names  # beyond 3 consecutive blanks — correctly dropped


# ── name_match_key / abbreviated name resolution ─────────────────────────────

def test_template_reader_parses_preferred_seats_before_name_column():
    from app.readers import template_reader
    from app.readers.template_reader import read_template

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.cell(row=1, column=2).value = template_reader._NAME_HEADER
    ws.cell(row=1, column=3).value = datetime.datetime(2026, 5, 4)
    ws.cell(row=2, column=1).value = "16,25; 624"
    ws.cell(row=2, column=2).value = "Ilya"

    path = _wb_to_path(wb)
    data = read_template(path, "Template")

    assert data.explicit_preferred_seats["Ilya"] == ["16.25", "624"]


def test_name_match_key_full_name():
    from app.utils.normalization import name_match_key
    assert name_match_key("Дононбаев Марат Болотович") == "до марат болотович"


def test_name_match_key_abbreviated():
    from app.utils.normalization import name_match_key
    assert name_match_key("До... Марат Болотович") == "до марат болотович"
    assert name_match_key("До Марат Болотович") == "до марат болотович"


def test_name_match_key_different_surnames_dont_collide():
    from app.utils.normalization import name_match_key
    # Дононбаев vs Дроздов — same 2 letters "др" would collide, but "до" vs "др" should not
    assert name_match_key("Дононбаев Марат Болотович") != name_match_key("Дроздов Марат Болотович")


def test_resolve_abbreviated_names_replaces_correctly():
    from app.domain.models import EmployeeDayChoice, EmployeeStatus
    from generate_seating import _resolve_abbreviated_names
    import datetime

    template_employees = ["Дононбаев Марат Болотович", "Драйзер Александр Михайлович"]
    choices = [
        EmployeeDayChoice("До... Марат Болотович", datetime.date(2026, 5, 4), EmployeeStatus.OFFICE),
        EmployeeDayChoice("Др... Александр Михайлович", datetime.date(2026, 5, 4), EmployeeStatus.REMOTE),
    ]
    resolved = _resolve_abbreviated_names(choices, template_employees)
    assert resolved[0].employee_name == "Дононбаев Марат Болотович"
    assert resolved[1].employee_name == "Драйзер Александр Михайлович"


def test_resolve_abbreviated_names_keeps_unknown_names():
    from app.domain.models import EmployeeDayChoice, EmployeeStatus
    from generate_seating import _resolve_abbreviated_names
    import datetime

    template_employees = ["Иванов Иван Иванович"]
    choices = [
        EmployeeDayChoice("Петров Пётр Петрович", datetime.date(2026, 5, 4), EmployeeStatus.OFFICE),
    ]
    resolved = _resolve_abbreviated_names(choices, template_employees)
    assert resolved[0].employee_name == "Петров Пётр Петрович"
