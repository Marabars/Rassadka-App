"""Smoke test: run full pipeline on the anonymized reference files."""
from __future__ import annotations

import sys
from pathlib import Path

import pytest

# Reference files relative to this test file
TESTS_DIR = Path(__file__).parent
REPO_DIR = TESTS_DIR.parent
SAMPLE_DIR = REPO_DIR.parent / "Rassadka v0.1"

CHOICES_FILE = SAMPLE_DIR / "График Май гугл Таблица Обезличенный.xlsx"
TEMPLATE_FILE = SAMPLE_DIR / "График Май Обезличенный.xlsx"
CONFIG_FILE = REPO_DIR / "config.yaml"


@pytest.mark.skipif(
    not CHOICES_FILE.exists() or not TEMPLATE_FILE.exists(),
    reason="Reference Excel files not found",
)
def test_smoke_full_pipeline(tmp_path):
    """Full pipeline produces a valid .xlsx with no crashes."""
    output = tmp_path / "output.xlsx"

    from generate_seating import run
    exit_code = run(
        choices_path=CHOICES_FILE,
        template_path=TEMPLATE_FILE,
        output_path=output,
        config_path=CONFIG_FILE,
    )

    assert output.exists(), "Output file was not created"
    assert output.stat().st_size > 0, "Output file is empty"

    import openpyxl
    wb = openpyxl.load_workbook(output, data_only=True)
    assert "Validation Report" in wb.sheetnames, "Validation Report sheet missing"

    # At least one main sheet exists
    main_sheets = [s for s in wb.sheetnames if s != "Validation Report"]
    assert main_sheets, "No main sheet found"

    ws = wb[main_sheets[0]]
    # Check that some cells have seat values assigned (not all empty)
    non_empty = sum(
        1
        for row in ws.iter_rows(min_row=6, values_only=True)
        for cell in row[2:]
        if cell is not None
    )
    assert non_empty > 0, "No seat assignments written to output"


@pytest.mark.skipif(
    not CHOICES_FILE.exists() or not TEMPLATE_FILE.exists(),
    reason="Reference Excel files not found",
)
def test_smoke_no_duplicate_seats_per_day(tmp_path):
    """No seat should be assigned to two employees on the same day."""
    output = tmp_path / "output_dup.xlsx"
    from generate_seating import run
    run(
        choices_path=CHOICES_FILE,
        template_path=TEMPLATE_FILE,
        output_path=output,
        config_path=CONFIG_FILE,
    )

    from app.readers.template_reader import read_template
    import yaml
    config = yaml.safe_load(CONFIG_FILE.read_text(encoding="utf-8"))
    template_sheet = config["input"]["template_sheet_name"]

    import openpyxl
    wb = openpyxl.load_workbook(output, data_only=True)
    ws = wb[template_sheet]

    from app.utils.excel_utils import find_header_row, collect_date_columns
    from app.utils.normalization import normalize_seat_id

    header_row = find_header_row(ws, 2, "ФИО")
    assert header_row is not None
    date_cols = collect_date_columns(ws, header_row, 3)

    for date, col_idx in date_cols.items():
        seats_this_day = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            raw = ws.cell(row=row_idx, column=2).value
            if raw is None:
                break
            seat = normalize_seat_id(ws.cell(row=row_idx, column=col_idx).value)
            if seat:
                seats_this_day.append(seat)
        assert len(seats_this_day) == len(set(seats_this_day)), (
            f"Duplicate seats on {date}: {seats_this_day}"
        )
