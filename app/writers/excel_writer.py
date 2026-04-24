from __future__ import annotations

import datetime
import shutil
from pathlib import Path
from typing import Optional

import openpyxl
import openpyxl.utils

from app.domain.models import GenerationResult
from app.readers.template_reader import TemplateData
from app.utils.excel_utils import copy_cell_style
from app.utils.normalization import seat_to_excel_value

_RESERVE_COUNT_LABEL = "резерв"


def write_output(
    template_path: Path,
    output_path: Path,
    template_data: TemplateData,
    result: GenerationResult,
    validation_sheet_name: str,
    add_new_employees: bool = True,
) -> None:
    """Copy template and append new date columns from choices. Historical columns are preserved."""
    shutil.copy2(template_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb[_get_main_sheet(wb)]

    # Build lookup: employee_name -> date -> seat_id
    assignment_lookup: dict[str, dict[datetime.date, Optional[str]]] = {}
    for asgn in result.assignments:
        assignment_lookup.setdefault(asgn.employee_name, {})[asgn.date] = asgn.seat_id

    # Find employees in choices that are not in the template
    new_employees = _new_employees_from_result(result, template_data.employee_order)

    existing_date_cols = template_data.date_cols
    first_emp_row = template_data.first_employee_row
    last_emp_row = template_data.last_employee_row

    # Insert rows for new employees BEFORE anything else (keeps row indices stable)
    row_shift = 0
    if add_new_employees and new_employees:
        insert_at = last_emp_row + 1
        ws.insert_rows(insert_at, amount=len(new_employees))
        row_shift = len(new_employees)
        for offset, name in enumerate(new_employees):
            ws.cell(row=insert_at + offset, column=2).value = name

    # Adjust reserve block rows after row insertion
    reserve_start = template_data.reserve_start_row + row_shift if template_data.reserve_start_row else 0
    reserve_end = template_data.reserve_end_row + row_shift if template_data.reserve_end_row else 0

    # Determine which dates from result are new (not already in the template)
    # Skip dates where nobody received a seat (all employees are day_off/remote/vacation)
    all_result_dates = {asgn.date for asgn in result.assignments}
    dates_with_office = {asgn.date for asgn in result.assignments if asgn.seat_id is not None}
    new_dates = sorted(
        d for d in all_result_dates
        if d not in existing_date_cols and d in dates_with_office
    )

    if not new_dates:
        _write_validation_sheet(wb, result, validation_sheet_name)
        wb.save(output_path)
        return

    # Add new date header columns to the sheet
    new_date_cols = _add_new_date_columns(
        ws, new_dates, existing_date_cols, template_data.header_row
    )

    # Write seat assignments for all employees in new date columns
    for offset, employee_name in enumerate(template_data.employee_order):
        row_idx = first_emp_row + offset
        _write_employee_seats(ws, row_idx, employee_name, new_date_cols, assignment_lookup)

    if add_new_employees and new_employees:
        for offset, employee_name in enumerate(new_employees):
            row_idx = last_emp_row + 1 + offset
            _write_employee_seats(ws, row_idx, employee_name, new_date_cols, assignment_lookup)

    # Write reserve seats and counts for new dates
    if reserve_start:
        _write_reserve_block(ws, reserve_start, reserve_end, new_date_cols, result.reserve_by_date)

    _update_reserve_count_row(ws, template_data.header_row, new_date_cols, result.reserve_by_date)

    _write_validation_sheet(wb, result, validation_sheet_name)
    wb.save(output_path)


def _add_new_date_columns(
    ws,
    new_dates: list[datetime.date],
    existing_date_cols: dict[datetime.date, int],
    header_row: int,
) -> dict[datetime.date, int]:
    """Append new date columns after the last existing column. Returns date -> col_idx mapping."""
    next_col = ws.max_column + 1
    last_existing_col = max(existing_date_cols.values()) if existing_date_cols else None

    new_date_cols: dict[datetime.date, int] = {}
    for date in new_dates:
        cell = ws.cell(row=header_row, column=next_col)
        cell.value = datetime.datetime(date.year, date.month, date.day)

        if last_existing_col:
            src_header = ws.cell(row=header_row, column=last_existing_col)
            copy_cell_style(src_header, cell)
            # Match column width
            src_letter = openpyxl.utils.get_column_letter(last_existing_col)
            dst_letter = openpyxl.utils.get_column_letter(next_col)
            src_width = ws.column_dimensions[src_letter].width
            if src_width:
                ws.column_dimensions[dst_letter].width = src_width

        new_date_cols[date] = next_col
        next_col += 1

    return new_date_cols


def _write_employee_seats(ws, row_idx, employee_name, date_cols, assignment_lookup):
    date_assignments = assignment_lookup.get(employee_name, {})
    for date, col_idx in date_cols.items():
        seat_id = date_assignments.get(date)
        if seat_id is not None:
            ws.cell(row=row_idx, column=col_idx).value = seat_to_excel_value(seat_id)


def _get_main_sheet(wb: openpyxl.Workbook) -> str:
    for name in wb.sheetnames:
        if "validation" not in name.lower():
            return name
    return wb.sheetnames[0]


def _new_employees_from_result(result: GenerationResult, template_order: list[str]) -> list[str]:
    """Return employees from result not in the template, in stable order."""
    template_set = set(template_order)
    seen: set[str] = set()
    extra: list[str] = []
    for asgn in result.assignments:
        name = asgn.employee_name
        if name not in template_set and name not in seen:
            seen.add(name)
            extra.append(name)
    return extra


def _write_reserve_block(
    ws,
    start_row: int,
    end_row: int,
    date_cols: dict[datetime.date, int],
    reserve_by_date: dict[datetime.date, list[str]],
) -> None:
    for row_idx in range(start_row, end_row + 1):
        for col_idx in date_cols.values():
            ws.cell(row=row_idx, column=col_idx).value = None

    for date, col_idx in date_cols.items():
        for offset, seat_id in enumerate(reserve_by_date.get(date, [])):
            row_idx = start_row + offset
            if row_idx > end_row:
                break
            ws.cell(row=row_idx, column=col_idx).value = seat_to_excel_value(seat_id)


def _update_reserve_count_row(
    ws,
    header_row: int,
    date_cols: dict[datetime.date, int],
    reserve_by_date: dict[datetime.date, list[str]],
) -> None:
    for row_idx in range(1, header_row):
        cell = ws.cell(row=row_idx, column=2)
        if cell.value is not None and str(cell.value).strip().lower() == _RESERVE_COUNT_LABEL:
            for date, col_idx in date_cols.items():
                ws.cell(row=row_idx, column=col_idx).value = len(reserve_by_date.get(date, []))
            return


def _write_validation_sheet(wb, result: GenerationResult, sheet_name: str) -> None:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(["Severity", "Date", "Employee", "Issue Code", "Description", "Suggested Action"])
    for issue in result.issues:
        ws.append([
            issue.severity.value,
            str(issue.date) if issue.date else "",
            issue.employee_name or "",
            issue.issue_code,
            issue.description,
            issue.suggested_action,
        ])
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
