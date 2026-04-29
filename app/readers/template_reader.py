from __future__ import annotations

import datetime
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from app.utils.excel_utils import collect_date_columns, find_header_row
from app.utils.normalization import normalize_employee_name, normalize_seat_id

_NAME_COL = 2
_NAME_HEADER = "ФИО"
_FIRST_DATA_COL = 3
_RESERVE_LABEL = "резерв мест"
_PREFERRED_SEAT_HEADER = "preferred seats"


@dataclass
class TemplateData:
    historical_assignments: dict[str, dict[datetime.date, str]] = field(default_factory=dict)
    all_seats: list[str] = field(default_factory=list)
    header_row: int = 5
    date_cols: dict[datetime.date, int] = field(default_factory=dict)
    first_employee_row: int = 6
    last_employee_row: int = 0
    reserve_start_row: int = 0
    reserve_end_row: int = 0
    employee_order: list[str] = field(default_factory=list)
    reserve_max_rows: int = 0
    explicit_preferred_seats: dict[str, list[str]] = field(default_factory=dict)


def read_template(path: Path, sheet_name: str) -> TemplateData:
    """Parse template seating file and extract structure + historical assignments."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]

    header_row = find_header_row(ws, _NAME_COL, _NAME_HEADER)
    if header_row is None:
        raise ValueError(
            f"Header row with '{_NAME_HEADER}' not found in first rows of "
            f"column {_NAME_COL} in sheet '{sheet_name}'. "
            "Check that the sheet name and column layout match config.yaml."
        )

    date_cols = collect_date_columns(ws, header_row, _FIRST_DATA_COL)
    first_employee_row = header_row + 1

    # Find optional explicit preferred seats column. Some templates name it
    # "Preferred Seats"; in the production layout it is also the column
    # immediately before the FIO column.
    preferred_seat_col: int | None = None
    for col in range(1, _FIRST_DATA_COL):
        val = ws.cell(row=header_row, column=col).value
        if val and str(val).strip().lower() == _PREFERRED_SEAT_HEADER:
            preferred_seat_col = col
            break
    if preferred_seat_col is None and _NAME_COL > 1:
        preferred_seat_col = _NAME_COL - 1

    # Single pass: classify each row below the header as employee, reserve, or skip.
    # Empty rows in the middle are skipped (real files often have blank separator rows).
    historical: dict[str, dict[datetime.date, str]] = {}
    explicit_preferred: dict[str, list[str]] = {}
    employee_order: list[str] = []
    last_employee_row = first_employee_row - 1
    reserve_start_row = 0
    reserve_end_row = 0

    for row_idx in range(first_employee_row, ws.max_row + 1):
        raw = ws.cell(row=row_idx, column=_NAME_COL).value
        if raw is None or str(raw).strip() == "":
            continue

        cell_text = str(raw).strip().lower()

        if cell_text == _RESERVE_LABEL:
            reserve_start_row = row_idx
            reserve_end_row = ws.max_row
            break  # everything below is reserve block

        employee_name = normalize_employee_name(str(raw))
        employee_order.append(employee_name)
        last_employee_row = row_idx

        if preferred_seat_col is not None:
            raw_pref = ws.cell(row=row_idx, column=preferred_seat_col).value
            if raw_pref is not None:
                seats = [
                    s for part in str(raw_pref).split(";")
                    if (s := normalize_seat_id(part.strip())) is not None
                ]
                if seats:
                    explicit_preferred[employee_name] = seats

        for date, col_idx in date_cols.items():
            raw_seat = ws.cell(row=row_idx, column=col_idx).value
            seat_id = normalize_seat_id(raw_seat)
            if seat_id is not None:
                historical.setdefault(employee_name, {})[date] = seat_id

    # Collect all seats from employee assignments + reserve block
    all_seats_set: set[str] = set()
    for seat_dict in historical.values():
        all_seats_set.update(seat_dict.values())

    if reserve_start_row:
        for row_idx in range(reserve_start_row, reserve_end_row + 1):
            for col_idx in range(_FIRST_DATA_COL, ws.max_column + 1):
                seat_id = normalize_seat_id(ws.cell(row=row_idx, column=col_idx).value)
                if seat_id is not None:
                    all_seats_set.add(seat_id)

    all_seats = sorted(all_seats_set, key=_seat_sort_key)
    reserve_max_rows = (reserve_end_row - reserve_start_row + 1) if reserve_start_row else 0

    wb.close()
    return TemplateData(
        historical_assignments=historical,
        all_seats=all_seats,
        header_row=header_row,
        date_cols=date_cols,
        first_employee_row=first_employee_row,
        last_employee_row=last_employee_row,
        reserve_start_row=reserve_start_row,
        reserve_end_row=reserve_end_row,
        employee_order=employee_order,
        reserve_max_rows=reserve_max_rows,
        explicit_preferred_seats=explicit_preferred,
    )


def _seat_sort_key(seat_id: str):
    try:
        return (0, float(seat_id))
    except ValueError:
        return (1, seat_id)
