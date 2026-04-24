from __future__ import annotations

import datetime
from pathlib import Path
from typing import Optional

import openpyxl

from app.domain.models import EmployeeDayChoice, EmployeeStatus, ValidationIssue
from app.domain.statuses import normalize_status
from app.domain.validation import unknown_status
from app.utils.excel_utils import collect_date_columns, find_header_row
from app.utils.normalization import normalize_employee_name

_NAME_COL = 1
_NAME_HEADER = "ФИО"
_FIRST_DATA_COL = 2
_MAX_CONSECUTIVE_EMPTY = 3  # stop after this many blank rows in a row


def read_choices(
    path: Path,
    sheet_name: str,
    status_mapping: dict[str, list[str]],
    target_dates: Optional[set[datetime.date]] = None,
) -> tuple[list[EmployeeDayChoice], list[ValidationIssue]]:
    """Read employee day choices from the Google Sheets export format."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name]

    header_row = find_header_row(ws, _NAME_COL, _NAME_HEADER)
    if header_row is None:
        raise ValueError(
            f"Header row with '{_NAME_HEADER}' not found in first rows of "
            f"column {_NAME_COL} in sheet '{sheet_name}'. "
            "Check that the sheet name and column layout match config.yaml."
        )

    date_cols = collect_date_columns(ws, header_row, _FIRST_DATA_COL)
    if target_dates:
        date_cols = {d: c for d, c in date_cols.items() if d in target_dates}

    choices: list[EmployeeDayChoice] = []
    issues: list[ValidationIssue] = []
    consecutive_empty = 0

    for row_idx in range(header_row + 1, ws.max_row + 1):
        raw_name = ws.cell(row=row_idx, column=_NAME_COL).value
        employee_name = normalize_employee_name(str(raw_name)) if raw_name is not None else ""

        if not employee_name:
            consecutive_empty += 1
            if consecutive_empty >= _MAX_CONSECUTIVE_EMPTY:
                break
            continue  # skip blank row, keep scanning

        consecutive_empty = 0

        for date, col_idx in date_cols.items():
            raw_status = ws.cell(row=row_idx, column=col_idx).value
            if raw_status is None or str(raw_status).strip() == "":
                status = EmployeeStatus.UNKNOWN
                raw_str = ""
            else:
                raw_str = str(raw_status).strip()
                status = normalize_status(raw_str, status_mapping)

            if status == EmployeeStatus.UNKNOWN and raw_str:
                issues.append(unknown_status(employee_name, date, raw_str))

            choices.append(EmployeeDayChoice(
                employee_name=employee_name,
                date=date,
                status=status,
            ))

    wb.close()
    return choices, issues
