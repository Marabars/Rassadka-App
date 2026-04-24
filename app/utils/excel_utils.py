from __future__ import annotations

import datetime
from typing import Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

_DATE_FORMATS = ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y")


def find_header_row(
    ws: Worksheet,
    name_col_index: int,
    name_header_value: str,
    max_scan_rows: int = 30,
) -> Optional[int]:
    """Return 1-based row index of the header row, or None if not found."""
    target = name_header_value.strip().lower()
    for row_idx in range(1, max_scan_rows + 1):
        cell = ws.cell(row=row_idx, column=name_col_index)
        if cell.value is not None and str(cell.value).strip().lower() == target:
            return row_idx
    return None


def collect_date_columns(
    ws: Worksheet,
    header_row: int,
    first_data_col: int,
) -> dict[datetime.date, int]:
    """Return mapping of date → 1-based column index from the header row.

    Handles datetime objects (standard Excel dates) and string-formatted dates
    (common when files are saved from Google Sheets or other tools).
    """
    date_cols: dict[datetime.date, int] = {}
    for col_idx in range(first_data_col, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val is None:
            continue
        if isinstance(val, datetime.datetime):
            date_cols[val.date()] = col_idx
        elif isinstance(val, datetime.date):
            date_cols[val] = col_idx
        elif isinstance(val, str):
            parsed = _try_parse_date(val.strip())
            if parsed is not None:
                date_cols[parsed] = col_idx
    return date_cols


def _try_parse_date(value: str) -> Optional[datetime.date]:
    for fmt in _DATE_FORMATS:
        try:
            return datetime.datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def copy_cell_style(src_cell, dst_cell) -> None:
    from copy import copy
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)
